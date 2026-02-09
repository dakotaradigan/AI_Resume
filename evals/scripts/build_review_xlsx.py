"""
Build an Excel file for human review of eval results.

Usage:
    python evals/scripts/build_review_xlsx.py

Reads the most recent eval run from evals/results/ and creates an .xlsx
with truncated responses, color-coded categories, and ready-to-fill grading columns.
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RESULTS_DIR = Path(__file__).parent.parent / "results"
DATASETS_DIR = Path(__file__).parent.parent / "datasets"

# Find most recent eval run
result_files = sorted(RESULTS_DIR.glob("eval_run_*.jsonl"))
if not result_files:
    print("No eval results found.")
    exit(1)

latest = result_files[-1]
print(f"Reading: {latest.name}")

results = []
with open(latest, "r", encoding="utf-8") as f:
    for line in f:
        if line.strip():
            results.append(json.loads(line))

MAX_RESPONSE_CHARS = 400

# Category colors
CATEGORY_FILLS = {
    "core": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "edge_case": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "unanswerable": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
    "off_topic": PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
    "adversarial": PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
}

wb = Workbook()
ws = wb.active
ws.title = "Eval Review"

# Headers
headers = [
    "ID", "Category", "Query", "Response (truncated)",
    "Response Time (s)", "Expected Topics",
    "PASS or FAIL", "Critique"
]
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="424242", end_color="424242", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Data rows
for i, r in enumerate(results, 2):
    response = r["response"]
    # Strip markdown bold markers for cleaner reading
    clean_response = response.replace("**", "").replace("##", "").replace("# ", "")
    truncated = clean_response[:MAX_RESPONSE_CHARS]
    if len(clean_response) > MAX_RESPONSE_CHARS:
        truncated += "..."

    row_data = [
        r["id"],
        r["category"],
        r["query"],
        truncated,
        r["response_time_s"],
        "; ".join(r.get("expected_topics", [])),
        "",  # pass/fail
        "",  # critique
    ]

    cat_fill = CATEGORY_FILLS.get(r["category"])
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=i, column=col, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = thin_border
        if cat_fill and col == 2:
            cell.fill = cat_fill

    # Highlight grading columns in yellow
    grade_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
    ws.cell(row=i, column=7).fill = grade_fill
    ws.cell(row=i, column=8).fill = grade_fill

# Column widths
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 45
ws.column_dimensions["D"].width = 60
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 40
ws.column_dimensions["G"].width = 14
ws.column_dimensions["H"].width = 35

# Freeze header row
ws.freeze_panes = "A2"

# Add rubric sheet
rubric = wb.create_sheet("Grading Rubric")
rubric_data = [
    ["Question to Ask", "What You're Checking", "If NO, Mark FAIL With:"],
    ["Is everything factually correct?", "Groundedness — no made-up info", 'FAIL: "hallucinated [what]"'],
    ["Did it answer what was asked?", "Correctness — addresses the question", 'FAIL: "didn\'t answer the question"'],
    ["Did it miss anything important?", "Completeness — covers relevant data", 'FAIL: "missed [what]"'],
    ["Is the tone right?", "Professional but personable", 'FAIL: "too casual / robotic"'],
    ["Is the length appropriate?", "Conciseness — right size for the question", 'FAIL: "too verbose"'],
    ["Did it handle boundaries well?", "For unanswerable/adversarial — stayed in lane", 'FAIL: "should have redirected"'],
]
for row_idx, row_data in enumerate(rubric_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = rubric.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.font = header_font_white
            cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
rubric.column_dimensions["A"].width = 35
rubric.column_dimensions["B"].width = 35
rubric.column_dimensions["C"].width = 35

output_path = DATASETS_DIR / "eval_review.xlsx"
wb.save(output_path)
print(f"Written {len(results)} rows to {output_path}")
print(f"\nOpen in Excel: {output_path}")
print(f"\nFill in columns G (PASS or FAIL) and H (Critique)")
print(f"See 'Grading Rubric' tab for guidance")

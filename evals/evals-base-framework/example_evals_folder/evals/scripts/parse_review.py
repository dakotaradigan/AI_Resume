"""
Parse the completed eval review Excel file and output a summary.

Usage:
    python evals/scripts/parse_review.py
"""

import json
from pathlib import Path
from openpyxl import load_workbook

DATASETS_DIR = Path(__file__).parent.parent / "datasets"
RESULTS_DIR = Path(__file__).parent.parent / "results"

xlsx_path = DATASETS_DIR / "eval_review.xlsx"
if not xlsx_path.exists():
    print(f"Review file not found: {xlsx_path}")
    exit(1)

wb = load_workbook(xlsx_path)
ws = wb["Eval Review"]

results = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    results.append({
        "id": row[0],
        "category": row[1],
        "query": row[2],
        "response_truncated": row[3],
        "response_time_s": row[4],
        "expected_topics": row[5],
        "human_label": str(row[6]).strip().lower() if row[6] else "",
        "human_critique": str(row[7]).strip() if row[7] else "",
    })

# Summary
total = len(results)
labeled = [r for r in results if r["human_label"] in ("pass", "fail")]
passes = [r for r in labeled if r["human_label"] == "pass"]
fails = [r for r in labeled if r["human_label"] == "fail"]
unlabeled = [r for r in results if r["human_label"] not in ("pass", "fail")]

print(f"{'='*60}")
print(f"EVAL REVIEW SUMMARY")
print(f"{'='*60}")
print(f"Total cases:  {total}")
print(f"Labeled:      {len(labeled)}")
print(f"  Pass:       {len(passes)} ({len(passes)/max(len(labeled),1)*100:.0f}%)")
print(f"  Fail:       {len(fails)} ({len(fails)/max(len(labeled),1)*100:.0f}%)")
if unlabeled:
    print(f"  Unlabeled:  {len(unlabeled)}")

# Pass/fail by category
print(f"\n{'='*60}")
print(f"RESULTS BY CATEGORY")
print(f"{'='*60}")
categories = sorted(set(r["category"] for r in labeled))
for cat in categories:
    cat_results = [r for r in labeled if r["category"] == cat]
    cat_pass = sum(1 for r in cat_results if r["human_label"] == "pass")
    cat_fail = sum(1 for r in cat_results if r["human_label"] == "fail")
    pass_rate = cat_pass / max(len(cat_results), 1) * 100
    print(f"  {cat:15s}  {cat_pass}/{len(cat_results)} pass ({pass_rate:.0f}%)  |  {cat_fail} fail")

# Failures detail
if fails:
    print(f"\n{'='*60}")
    print(f"FAILURE DETAILS")
    print(f"{'='*60}")
    for r in fails:
        print(f"\n  [{r['id']}] ({r['category']}) {r['query'][:60]}")
        print(f"      Response time: {r['response_time_s']}s")
        if r["human_critique"]:
            print(f"      Critique: {r['human_critique']}")

# Latency summary
print(f"\n{'='*60}")
print(f"LATENCY SUMMARY")
print(f"{'='*60}")
times = [r["response_time_s"] for r in results if r["response_time_s"]]
if times:
    times_sorted = sorted(times)
    avg = sum(times) / len(times)
    p50 = times_sorted[len(times_sorted)//2]
    p95 = times_sorted[int(len(times_sorted)*0.95)]
    over_3s = sum(1 for t in times if t > 3.0)
    print(f"  Average:  {avg:.2f}s")
    print(f"  Median:   {p50:.2f}s")
    print(f"  P95:      {p95:.2f}s")
    print(f"  Over 3s:  {over_3s}/{len(times)} ({over_3s/len(times)*100:.0f}%)")

# Critique patterns (for failure mode grouping)
if fails:
    print(f"\n{'='*60}")
    print(f"CRITIQUE PATTERNS (for failure mode table)")
    print(f"{'='*60}")
    critiques = {}
    for r in fails:
        c = r["human_critique"] or "(no critique)"
        critiques.setdefault(c, []).append(r["id"])
    for critique, ids in sorted(critiques.items(), key=lambda x: -len(x[1])):
        print(f"  [{len(ids)}x] {critique}")
        print(f"       IDs: {ids}")

# Save parsed results as JSONL for downstream use
output_path = RESULTS_DIR / "human_labeled_results.jsonl"
with open(output_path, "w", encoding="utf-8") as f:
    for r in results:
        f.write(json.dumps(r) + "\n")
print(f"\nParsed results saved to: {output_path}")
